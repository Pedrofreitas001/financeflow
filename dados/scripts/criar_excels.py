import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def criar_excel_cash_flow():
    with open('dados_cash_flow_exemplo.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    df = pd.DataFrame(data)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fluxo de Caixa"
    
    # Escrever headers
    headers = ['ID', 'Mês', 'Empresa', 'Tipo', 'Categoria', 'Data Vencimento', 'Valor', 'Status', 'Responsável']
    ws.append(headers)
    
    # Formatar headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever dados
    for item in data:
        ws.append([
            item.get('id', ''),
            item.get('mes', ''),
            item.get('empresa', ''),
            item.get('tipo', ''),
            item.get('categoria', ''),
            item.get('data_vencimento', ''),
            item.get('valor', 0),
            item.get('status', ''),
            item.get('responsavel', '')
        ])
    
    # Ajustar largura de colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    
    wb.save('../excel_exemplos/CashFlow_Exemplo.xlsx')
    print("✅ CashFlow_Exemplo.xlsx criado")

def criar_excel_indicadores():
    with open('dados_indicadores_exemplo.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    df = pd.DataFrame(data)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"
    
    # Escrever headers
    headers = ['Mês', 'Empresa', 'ROE %', 'ROA %', 'Margem Líquida %', 'Margem Operacional %', 
               'Liquidez Corrente', 'Liquidez Seca', 'Endividamento %', 'Alavancagem', 
               'Giro Ativo', 'Prazo Recebimento', 'Prazo Pagamento']
    ws.append(headers)
    
    # Formatar headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever dados
    for item in data:
        ws.append([
            item.get('mes', ''),
            item.get('empresa', ''),
            round(item.get('roe', 0), 2),
            round(item.get('roa', 0), 2),
            round(item.get('margemLiquida', 0), 2),
            round(item.get('margemOperacional', 0), 2),
            round(item.get('liquidezCorrente', 0), 2),
            round(item.get('liquidezSeca', 0), 2),
            round(item.get('endividamento', 0), 2),
            round(item.get('alavancagem', 0), 2),
            round(item.get('giroAtivo', 0), 2),
            item.get('prazoRecebimento', 0),
            item.get('prazoPagamento', 0)
        ])
    
    # Ajustar largura de colunas
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i)].width = 16
    
    wb.save('../excel_exemplos/Indicadores_Exemplo.xlsx')
    print("✅ Indicadores_Exemplo.xlsx criado")

def criar_excel_orcamento():
    with open('dados_orcamento_exemplo.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    df = pd.DataFrame(data)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"
    
    # Escrever headers
    headers = ['Mês', 'Empresa', 'Categoria', 'Orçado', 'Realizado', 'Variância', 'Variância %', 'Responsável', 'Observações']
    ws.append(headers)
    
    # Formatar headers
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever dados
    for item in data:
        orcado = item.get('orcado', 0)
        realizado = item.get('realizado', 0)
        variancia = realizado - orcado
        variancia_pct = (variancia / orcado * 100) if orcado > 0 else 0
        
        ws.append([
            item.get('mes', ''),
            item.get('empresa', ''),
            item.get('categoria', ''),
            orcado,
            realizado,
            variancia,
            round(variancia_pct, 2),
            item.get('responsavel', ''),
            item.get('observacoes', '')
        ])
    
    # Ajustar largura de colunas
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i)].width = 15
    
    wb.save('../excel_exemplos/Orcamento_Exemplo.xlsx')
    print("✅ Orcamento_Exemplo.xlsx criado")

def criar_excel_despesas():
    with open('dados_despesas_exemplo.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    df = pd.DataFrame(data)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Despesas"
    
    # Escrever headers
    headers = ['Ano', 'Mês', 'Empresa', 'Categoria', 'Subcategoria', 'Valor']
    ws.append(headers)
    
    # Formatar headers
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever dados
    for item in data:
        ws.append([
            item.get('ano', ''),
            item.get('mes', ''),
            item.get('empresa', ''),
            item.get('categoria', ''),
            item.get('subcategoria', ''),
            item.get('valor', 0)
        ])
    
    # Ajustar largura de colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    
    wb.save('../excel_exemplos/Despesas_Exemplo.xlsx')
    print("✅ Despesas_Exemplo.xlsx criado")

# Gerar os 4 arquivos
criar_excel_cash_flow()
criar_excel_indicadores()
criar_excel_orcamento()
criar_excel_despesas()
print("\n✅ Todos os arquivos Excel foram criados com sucesso!")
